import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image

def insert_image_centered(file_path, image_path, sheet_name='Sheet1', cell='A1'):
    # 檢查圖片文件是否存在
    if not os.path.isfile(image_path):
        raise FileNotFoundError(f"Image file not found: {image_path}")

    # 加載或創建工作簿
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()  # 如果文件不存在，則創建新文件

    # 獲取工作表
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)

    # 創建圖片對象
    img = Image(image_path)

    # 獲取儲存格的行高和列寬
    cell_row = int(cell[1:])  # 行數（例如 A1 -> 1）
    cell_col = cell[0]  # 列字母（例如 A1 -> A）
    
    # 轉換列字母為列數（例如 A -> 1, B -> 2, ...）
    col_idx = ord(cell_col.upper()) - ord('A') + 1
    
    # 獲取儲存格的尺寸
    row_height = sheet.row_dimensions[cell_row].height or 15  # 預設行高
    col_width = sheet.column_dimensions[cell_col].width or 8.43  # 預設列寬

    # 計算圖片大小
    img_width = img.width
    img_height = img.height
    
    # 計算需要縮放的比例
    scale_width = col_width * 180  # 1個Excel列的寬度約等於7.5個像素
    scale_height = row_height * 180  # 1個Excel行的高度約等於0.75個像素

    scale_ratio = min(scale_width / img_width, scale_height / img_height)

    # 根據比例縮放圖片
    img.width = int(img_width * scale_ratio)
    img.height = int(img_height * scale_ratio)

    # 計算圖片的左上角位置
    left = (col_idx - 1) * 7.5  # 每列大約7.5像素
    top = (cell_row - 1) * 15  # 每行大約15像素

    # 設置圖片的絕對位置
    img.anchor = cell  # 將圖片放置在儲存格中
    img.left = left + (col_width * 7.5 - img.width) / 2
    img.top = top + (row_height * 0.75 - img.height) / 2

    # 插入圖片到指定單元格
    sheet.add_image(img)

    # 保存工作簿
    workbook.save(file_path)

# 使用示例
if __name__ == "__main__":
    excel_file = 'example.xlsx'  # 指定 Excel 文件的名稱
    image_file = './image/logo.png'  # 指定要插入的圖片路徑
    insert_image_centered(excel_file, image_file)
